
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import date
from openpyxl import Workbook
from openpyxl import load_workbook
import time

# 크롤링할 사이트 및 로딩 대기
driver = webdriver.Chrome()
driver.get("https://www.naver.com/")
time.sleep(2)

# 검색창 선택, 검색어 선택, 엔터치는 것, 로딩 대기
elem = driver.find_element_by_name("query")
elem.send_keys("Fed")
elem.send_keys(Keys.RETURN)
time.sleep(2)

# 검색 결과가 나온 창에서 최신순(a herf) 를 선택
driver.find_element_by_partial_link_text('최신순').click()
# 또는 driver.find_element_by_link_text('Send InMail').click() 도 사용 가능
time.sleep(2)

# 뉴스 타이틀 수집
news_titles = driver.find_elements_by_css_selector(".news_tit")
for i in news_titles:
    title = i.text
    print(title) 
# 뉴스 하이퍼링크 수집
for i in news_titles:
    href = i.get_attribute('href')
    print(href) 

#창 넘기기


# 엑셀파일 불러오기
wb = load_workbook("news2202.xlsx")
# 날짜 이름으로 된 시트 생성
today = date.today()
ws_want = wb.create_sheet(today, index=0)
ws = wb["today"]


# 파일 생성 및 저장
wb = Workbook( ) # 엑셀파일 생성
wb.save("test.xlsx") # 엑셀 파일 저장
ws = wb.active # 첫번째 시트 지정
ws_new = wb.create_sheet() # 새로운 시트 추가
ws.title = "First Sheet" # 시트이름 바꾸는 방법
ws_new.title = "New Sheet"
ws_want = wb.create_sheet("Third sheet", 1) # 새로운 시트를 특정위치에 생성할때 1은 두번째에 생성됨
ws.sheet_properties.tabColor = "f4566e" # 시트 색깔 변경 RGB컬러코드이용

# 엑셀 파일 불러오기
wb = load_workbook("test.xlsx") # 엑셀파일 열기
ws = wb["sheetname"] # 사용할 시트 선택 필수!! 꼭 지정해줘야함
    
# 셀에 입력하기
# A1 셀에 값 입력
ws["A1"] = 5
# B1 셀에 값 입력 (추천하는 방법)
ws.cell(row=1, column=2, value=70)
# 이렇게도 가능합니다
ws.cell(1, 2, value=70)

# 셀 내용 삭제
ws["A1"] = ""

# 행 삭제
# 1번행 1개 삭제
ws.delete_rows(1)
# 1번행부터 총 3개의 행 삭제 (1,2,3행 삭제)
ws.delete_rows(1, 3)

# 열 삭제
# 3번열 1개 삭제
ws.delete_cols(3)
# 3번열부터 총 3개의 열 삭제 (3,4,5열 삭제)
ws.delete_cols(3, 3)

# 빈 행 삽입
# 2번행에 빈행 삽입
ws.insert_rows(2)
# 1번행부터 총 3개의 빈행 삽입
ws.insert_rows(1, 3)

# 빈 열 삽입
# 3번열에 빈열 삽입
ws.insert_cols(3)
# 3번열부터 총 3개의 빈열 삽입
ws.insert_cols(3, 3)

# 셀 내용 잘라서 다른 셀에 붙여넣기
#D1 셀의 내용을 오른쪽으로 1열 이동시킴
ws.move_range("D1", cols=2)
#D4부터 F10 영역의 데이터를 위쪽으로 1행, 오른쪽으로 2열 이동시킴
ws.move_range("D4:F10", rows=-1, cols=2)

#시트 선택하기

test = wb.get_sheet_by_name("이름")   # '이름'이라는 워크시트 불러오기
test = wb.get_active_sheet()  # 활성시트 가져오기, 활성시트란 엑셀파일을 열었을 때 default로 열리는 시트이다.

