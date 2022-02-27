from datetime import date
from openpyxl import load_workbook


# 엑셀파일 불러오기
wb = load_workbook("new2202.xlsx")
# 날짜 이름으로 된 시트 생성
todaynow = date.today()
ws_want = wb.create_sheet(todaynow.strftime('%Y-%m-%d'), index=0)
wb.save("new2202.xlsx")
# 입력할 시트 선택
ws = wb.active
# 제목 입력
ws["A1"] = "제목"
ws["B1"] = "링크"
wb.save("new2202.xlsx")

#데이터 입력
ws.cell(2,1).value = '뉴시스'
ws.cell(2,2).value = 'add'
wb.save("new2202.xlsx")

